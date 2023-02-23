import os
from openpyxl import load_workbook

# set the directory path to your main folder
directory = '/Users/BenBurkert/Library/CloudStorage/OneDrive-bwedu/Studium/OSU/Winter Term/BIS 371/research_productivity'

# loop through each folder in the directory
for foldername in os.listdir(directory):
    data = {}
    # check if the current item in the directory is a folder
    if os.path.isdir(os.path.join(directory, foldername)):
        # set the path to the Excel file in the current folder
        excel_path = os.path.join(directory, foldername, foldername + '.xlsx')
        # load the workbook
        wb = load_workbook(excel_path)
        # select the worksheet
        ws = wb.worksheets[0]
        # Loop through the rows
        for row in ws.iter_rows(min_row=7, values_only=True):
            data = {
                "paper_title": row[0],
                "paper_type": row[0]
            }
            print(data)
