from openpyxl import load_workbook


class DataExtractor:

    def __init__(self, file_path):
        self._file_path = file_path

    def get_worksheet(self):
        """Returns the worksheet in the Excel file in which the data is contained."""
        wb = load_workbook(self._file_path)
        ws = wb.worksheets[0]
        return ws

    def get_paper_data(self, worksheet, row):
        """Retrieving the data related to one paper."""
        paper_data = {
            "paper_title": worksheet["A" + str(row)].value,
            "target_type": worksheet["B" + str(row)].value,
            "target": worksheet["C" + str(row)].value,
            "tier": worksheet["D" + str(row)].value,
            "role": worksheet["J" + str(row)].value,
        }
        return paper_data

    def get_coauthor_data(self, worksheet, row):
        """Retrieving all the coauthors for one paper."""
        coauthor_data = {
            "co_author_1": worksheet["E" + str(row)].value,
            "co_author_2": worksheet["F" + str(row)].value,
            "co_author_3": worksheet["G" + str(row)].value,
            "co_author_4": worksheet["H" + str(row)].value,
        }
        return coauthor_data

    def get_activity_data(self, worksheet, row):
        """Retrieving all the activities related to one paper."""
        activity_data = {
            "activity_1": [worksheet["K" + str(row)].value, worksheet["L" + str(row)].value]
        }
        counter = 1
        maximum_row = worksheet.max_row
        # as long as no new paper is listed in the file, add activities to dictionary
        while worksheet["A" + str(row + counter)].value is None:
            # check if the row number is higher than the last one in the sheet with values, if yes: stop the loop
            if (row + counter) <= maximum_row:
                key = f"activity_{counter + 1}"
                value = [worksheet["K" + str(row + counter)].value, worksheet["L" + str(row + counter)].value]
                activity_data[key] = value
                counter += 1
            else:
                break
        return activity_data

    def get_faculty_department_data(self, worksheet):
        """Retrieving the data for the department in which the faculty member is in."""
        names = worksheet["A3"].value.split(' ')
        faculty_department_data = {
            "f_name": names[0],
            "l_name": names[1],
            "department": worksheet["B3"].value
        }
        return faculty_department_data
